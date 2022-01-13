# -*- coding: utf-8 -*-
from PyQt5.QtWebEngineWidgets import QWebEngineView
# from kiwipiepy import Kiwi
import codecs, random
from openpyxl import load_workbook
from tqdm import tqdm
# from seqeval.metrics.sequence_labeling import weighted_score
# from pred import *
# from stats import *
from PyQt5 import uic
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
## From MALA
# import convert, Eval_def
from KoBERT_NER import kobert_main
from KoBERT_NER.predict import predict
from KoBERT_NER.utils import MODEL_PATH_MAP, MODEL_CLASSES, init_logger
# from linitologo_rc import *

## 조사 제외를 위한 형태소 분석 패키지
from koalanlp.Util import initialize, finalize
from koalanlp import API
from koalanlp.proc import SentenceSplitter, Tagger, Parser, Dictionary

import argparse
import csv, sys, os, re

# from trainer import Trainer
# from utils import init_logger, load_tokenizer, set_seed, MODEL_CLASSES, MODEL_PATH_MAP
# from data_loader import load_and_cache_examples

import subprocess
def file_start(filename):
    if sys.platform == "win32":
        os.startfile(filename)
    else:
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener,filename])

# def resource_path(relative_path):
#     """ Get absolute path to resource, works for dev and for PyInstaller """
#     base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
#     return os.path.join(base_path, relative_path)
#
# form = resource_path("./UI/LIEN.ui")
# form_class = uic.loadUiType (form)[0]
form_class = uic.loadUiType ("./UI/LISA.ui")[0]


def file_start(filename):
    if sys.platform == "win32":
        os.startfile(filename)
    else:
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener,filename])

def make_color(num):

    # hex color의 재료들을 랜덤 생성하는 리스트 (어두운 계열 피하게끔 설정)
    hex_color_list = [chr (c) for c in range (ord ('a'), ord ('f') + 1)] + [str (n) for n in range (8, 10)]

    # hex_color_list의 원소를 랜덤하게 추출할 변수
    loc = random.randint (0, len (hex_color_list) - 1)

    # 처음 들어왔을 때 hex color를 추출하기 위해 앞에 #을 붙여준다
    if num == 6:
        return "#" + hex_color_list[loc] + make_color (num - 1)

    # 6번을 돌게 되면 재귀함수를 종료시킨다
    if num == 1:
        return hex_color_list[loc]

    # hex color를 만드는 재귀함수 호춫
    return hex_color_list[loc] + make_color (num - 1)


def check_color(color_dic):
    # check color 함수가 실행되면 color에 만들어진 색깔을 저장
    color = make_color (6)

    if color in color_dic.values() or color == '000000' or color == 'ffffff' or color == "#d3d3d3":
        check_color ()
    else:
        return color

def hex2rgb(hex_code):
    hex = hex_code.lstrip("#")
    return "rgba" + str(tuple(int(hex[i:i+2], 16) for i in (0, 2, 4))).rstrip(")")

class ProcessThread (QThread):
    result_txt = ''
    finish_sig = pyqtSignal(int)

    def __init__(self, parent=None):
        super ().__init__ ()

    def run(self):
        global vislist, result_txt
        finint = 0
        self.finish_sig.emit(finint)
        color_path = os.getcwd () + r'/DefaultDic.xlsx'
        ### HTML 시작-끝-띄어쓰기단위 처리를 위한 스트링
        sent_start = """<!DOCTYPE html>
<html>
  <head>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR&family=Outfit:wght@300;700&display=swap" rel="stylesheet">
    <link rel="stylesheet" type="text/css" href="../textstyle.css">
    
    <script src="../click_functions.js">/* java script 파일 연결 */</script>

    <title>DecoNERO</title>
  </head>
    
  <body>
    <h1 style="font-family: 'Outfit', sans-serif; font-weight: 700"><img src="../UI/DICORA-logo.png" alt="DICORA Logo" style="width:2.0rem;height:2.0rem; margin-right:0.5rem;">DecoNERO</h1>
    <div class='groove'>
      <h2 style="font-family: 'Outfit', sans-serif; font-weight: 700; text-indent: 0.9rem; margin-top: 1.2rem; margin-bottom: 0.4rem">TAG COLOR</h2>
      <input type="checkbox" id='showall' onclick='getCheckboxValue(event)'>
      <label for="showall" class="showall">show all</label><br>

      %s
      </div>
    <div class='a'>"""
        sent_end = """</div>\n</body>\n</html>"""

        js_file = """"""

        js_checkbox="""function getCheckboxValue(event) {
        var ptag2 = document.querySelectorAll('.entity');
        var ptag = document.querySelectorAll(".tagset");
        
        var re = /(rgba\(\d+\, \d+\, \d+\, )0(\))/;
        var re2 = /rgb(\(\d+\, \d+\, \d+)(\))/;
        
        for(var i=0;i<ptag2.length;i++){

            let result = '';
            if(event.target.checked)  {
            var bc = ptag2[i].style.backgroundColor;
            var newbc = bc.replace(re, '$11$2');
            ptag2[i].style.backgroundColor=newbc;
            ptag[i].style.display = 'inline';
            }else {
            var bc = ptag2[i].style.backgroundColor;
            var orgbc = bc.replace(re2, 'rgba$1\, 0$2');
            ptag2[i].style.backgroundColor=orgbc;
            ptag[i].style.display = 'none';
            }
        }
        }
"""
        js_function = """\nfunction %sent() {
		var ptag2 = document.querySelectorAll('%sent');
        var ptag = document.querySelectorAll("%s");
        var re = /(rgba\(\d+\, \d+\, \d+\, )0(\))/;
        var re2 = /rgb(\(\d+\, \d+\, \d+)(\))/;

    for(var i=0;i<ptag2.length;i++){

      if (ptag2[i].style.backgroundColor == "%s, 0)") {
        var bc = ptag2[i].style.backgroundColor;
      	var newbc = bc.replace(re, '$11$2');
        ptag2[i].style.backgroundColor=newbc;
        ptag[i].style.display = 'inline';
      } else {
        var bc = ptag2[i].style.backgroundColor;
      	var orgbc = bc.replace(re2, 'rgba$1\, 0$2');
        ptag2[i].style.backgroundColor=orgbc;
        ptag[i].style.display = 'none';
      }
    }
	}
    """

        ## Tag Color 보여주기
        css_tag_color = """<button class="taginfo" style="background-color: {color};" onclick="{id}ent();">{tag}</button>"""

        ### 시각화 마크
        markup = """<{id}ent id="{id}" class="entity" style="background: {entcolor}, 0)">{ent}</{id}ent><{id} id="{id}TAG" class="tagset">{tag}</{id}>"""
        # markup = """<mark class="entity" style="color:#FFFFFF; background: %s; padding: 0.45em 0.6em; line-height: 2.5; margin: 0 0.25em; border-radius: 0.35em;">%s<span style="color:#000000; background:#FFFFFF; padding: 0.45em 0.6em; font-size: 0.7em; font-weight: bold; border-radius: 1em; text-transform: uppercase; vertical-align: middle; margin-left: 0.5rem">%s</span></mark>"""


        color_file = load_workbook (color_path)
        # 태그 사전 받기
        color_sheet = color_file[mainui.Tagsetmod_combo.currentText()]

        color_dic = {}
        color_fin = {}        
        css_tag_info = """"""

        # 파스텔 톤 헥스 코드 62 개
        pastel_hex_list = [
            "#BFC8D7", '#E2D2D2', '#A2B59F', '#E3E2B4', '#E8E7D2', '#D2D5B8', '#BDC2BB', '#C9BA9B', '#F0E4D4', '#F9D9CA', '#D18063', '#917B56',
            '#EDE1E3', '#D1DFE8', '#909FA6', '#FADCDA', '#EEB8B8', '#C5DAD1', '#D5CB8E', '#ECD4D4', '#CCDBE2', '#C9CBEO', '#F8DAE2', '#DEB3CF',
            '#B57FB3', '#EADB80', '#AEDDEF', '#E1B4D3', '#FAF0E4', '#EECFBB', '#F6B99D', '#CB8A90', '#EACACB', '#E2B3A3', '#A3B6C5', '#B1D3C5',
            '#CFDD8E', '#E4BEB3', '#FCE9DA', '#FFCEC7', '#FFD0A6', '#E098AE', '#E9E1D4', '#FSDDAD', '#F1BCAE', '#C9DECF', '#E9E0CF', '#E9CEB9',
            '#E5BFBC', '#B97687', '#C6D2BE', '#83B1C9', '#E5C1C5', '#C3E2DD', '#6ECEDA', '#CACFE3', '#838BB2', '#E4A99B', '#CE8467', '#C7D6D8',
            '#FFD6AA', '#EFBAD6'
            ]

        for tokens in color_sheet.rows:
            temp = []
            for cell in tokens:
                temp.append (str (cell.value).strip ())
            if temp == ["Tag", "Mean", "Color", "Visualize"]:
                pass
            else:
                ## visualize에 1이라고 표시된 것만 시각화하기
                if temp[-1] == '1':
                    color_dic[temp[0]] = temp[2]
                    color_fin[temp[0]] = temp[1]

                    # Visualize 하기로 한 것중 color가 부여 안된 항목에 대해 color 랜덤 부여
                    if color_dic[temp[0]] == 'None':
                        
                        # 우선적으로 파스텔 톤 지정색을 부여한다.
                        if len(pastel_hex_list) != 0:
                            color_dic[temp[0]] = pastel_hex_list.pop()

                        # 파스텔 톤 색을 다 지정하면  부여된 색과 흰색, 검정색, 옅은 회색을 제외하고 색을 랜덤하게 만들어 부여한다.
                        else: color_dic[temp[0]] = check_color(color_dic)
                    
                    ## html css taginfo 부분 만들기
                    css_tag_info += css_tag_color.format(color=color_dic[temp[0]], id=temp[1], tag=temp[1])

                    ## java_script file 만들기
                    js_file += js_function % (temp[1], temp[1], temp[1], hex2rgb(color_dic[temp[0]]))
                else:
                    pass
        js_file += js_checkbox
        color_file.close ()

        ## java_script file 저장하기
        js_file_path = os.getcwd() + r'/click_functions.js'
        with codecs.open (js_file_path, 'w', encoding='utf-8') as f:  # Use some reasonable temp name
            f.write (js_file)
            f.close ()

        my_entity = [i for i in color_dic]
        target_path = mainui.targetInput.text ()
        # LISA랑 LIEN은 항상 모델 사용
        # if mainui.koBERTCheck.isChecked () == True:
        model_path = mainui.modelPath.text ()
        target_text, _ = predict (pred_model=model_path, pred_input=target_path, pred_label=model_path + "/label.txt",  pred_batch_size=int(64))
        # 형태소 분석 및 조사 후치 기능 사용 검사
        if mainui.Use_Declension_Check.isChecked() == True:
            initialize(OKT='LATEST')
            if mainui.Dict_input.text().strip() != '':
                self.Add_User_Dict()
            tagger = Tagger(API.OKT)
            target_text = self.Check_Josa_and_Replace(target_text)
        # else:
        #     target_text = codecs.open (target_path, 'r', encoding='utf-8').read ()
        #     target_text = target_text.replace ('{S}', '\n')
        
        target_text = re.sub(r'[\r\n]{2,}',r'\n', target_text)

        target_text = target_text.split ('\n')
        entity_dic = []
        mid = r'[^\n\<]+'
        ent_reg = re.compile (r'<.+?>' + mid + r'</.+?>')
        htm_linelis = []
        for i in range (len (target_text)):
            line = target_text[i]
            htm_line = target_text[i]
            if re.search (ent_reg, line) != None:
                pair = re.findall (r'<(.+?)>(' + mid + r')</.+?>', line)
                repl = re.findall (r'<.+?>' + mid + r'</.+?>', line)

                ent_tag_pair = []
                ent_repl_list = []
                
                # 왜 인지는 모르겠지만 pair에 빈 튜플 ('', '')이 발생해 길이 오류가 있어 제거해주는 코드
                pair = [sets for sets in pair if sets != ('', '')]
                
                
                # 정규표현식으로 찾아낸 정보들이 입력한 Entity나 Entities에 해당하는 경우에만
                # ent_tag_pair 리스트와 ent_repl_list에 정보들을 추가해준다
                for seq in range (len (pair)):
                    if pair[seq][0] in my_entity:
                        ent_tag_pair.append (pair[seq])
                        ent_repl_list.append (repl[seq])

                counter = -1
                if ent_tag_pair not in [[], None]:
                    for tagset, entity in ent_tag_pair:
                        counter += 1
                        
                        htm_line = htm_line.replace (ent_repl_list[counter],
                                                     markup.format(id = color_fin[tagset], entcolor = hex2rgb(color_dic[tagset]), ent = entity, tag = color_fin[tagset]))
                        line = line.replace (ent_repl_list[counter],
                                             '<' + tagset + '>' + entity + '</' + tagset + '>')
                        entity_dic.append ([entity, [tagset, color_dic[tagset]]])

                    target_text[i] = line
                    htm_linelis.append ('<p>' + htm_line + '</p>' )
                else:
                    htm_linelis.append ('<p>' + htm_line + '</p>')
                    pass
            else:
                htm_linelis.append ('<p>' + htm_line + '</p>')
                pass
            
        result_htm = '\n'.join (htm_linelis)
        result_txt = '{S}\n'.join (target_text)

        result_htm = sent_start % (css_tag_info) + re.sub ('({S}\n)+', '', result_htm) + sent_end
        Textvissaveloc = os.getcwd () + r'/Htmls/Textvis.html'
        with codecs.open (Textvissaveloc, 'w', encoding='utf-8') as f:  # Use some reasonable temp name
            f.write (result_htm)
            f.close ()
            
        # 결과 txt 파일로 자동 저장        
        result_txt_path = os.getcwd() + r'/LISA_result_text.txt'
        with codecs.open (result_txt_path, 'w', encoding='utf-8') as f:  # Use some reasonable temp name
            f.write (result_txt)
            f.close ()

        # 바로 결과 보여주기
        file_start(result_txt_path)

        # 0101 사이즈 조절 추가
        # 0113 통계 그래프 사용 안함    
        # vislist = Statistic_deco (result_txt, ent_dic=entity_dic, uicol=color_dic)
        # vislist[0] = Textvissaveloc
        finint = 1
        self.finish_sig.emit (finint)

     # 형태소 분석 사용자 사전 추가 펑션
    def Add_User_Dict(self):
        Kkma_Dict = Dictionary(API.OKT)

        User_Dicts = self.FSPairing_Dictfile.text().strip('"').split('", "')
        for dic in User_Dicts:
            dict_df = pd.read_excel(dic, header=None, sheet_name=None)
            sheet_names = list(dict_df.keys())

            tag_mapper = {'NS': POS.NNG, 'DS': POS.MAG, 'VS': POS.VV, 'AS': POS.VA}
            lexeme_tag_ls = []

            for s in sheet_names:
                temp = list(zip(dict_df[s][0], dict_df[s][1]))
                lexeme_tag_ls += list(map(lambda x: (x[0], tag_mapper[x[1]]), temp))

            for lex in lexeme_tag_ls:
                Kkma_Dict.addUserDictionary(lex)

    # 조사 체크 및 처리를 위한 펑션
    def Check_Josa_and_Replace(self, TargetText):
        # initialize(OKT='LATEST')

        tagger = Tagger(API.OKT)

        TempText = TargetText.split('\n')
        exp = []
        for text in tqdm(TargetText, desc='Searching Expressions'):
            expression = re.findall('([^<>]+?)(</.+?>)[$\.\!?\s]', text)
            for e in expression:
                temp = '%s%s%s' % (e[1].replace('</', '<', 1), e[0], e[1])
                if temp in text:
                    if e not in exp:
                        exp.append(e)
        original_list = []
        replace_list = []
        for toks in tqdm(exp, desc='Checking Josa'):
            check_tok = toks[0]
            tag_open = toks[1].replace('</', '<')
            tag_close = toks[1]

            sentence = tagger(check_tok)
            for sent in sentence:
                replace_text = ''
                if len(sent) > 1:
                    for postword in sent[:-1]:
                        replace_text += postword.getSurface()
                word = sent[-1]
                analyzed_tok = []
                for morph in word:
                    analyzed_tok.append("%s\t%s " % (morph.getSurface(), morph.getTag()))
                josa_exist = False
                for tok in analyzed_tok:
                    if 'JX' not in tok.split('\t')[1]:
                        replace_text += tok.split('\t')[0]
                    else:
                        if josa_exist == False:
                            replace_text += tag_close + tok.split('\t')[0]
                            josa_exist = True
                        else:
                            replace_text += tok.split('\t')[0]
                if josa_exist != False:
                    original_text = check_tok + tag_close
                    if ' ' in original_text:
                        flg = 0
                        for i in range(original_text.count(' ')):
                            space = original_text.find(' ', flg)
                            replace_text = replace_text[:space] + ' ' + replace_text[space:]
                            flg = space + 1

                    original_text = tag_open + original_text
                    replace_text = tag_open + replace_text

                    original_list.append(original_text)
                    replace_list.append(replace_text)

        finalize()

        for o, r in zip(original_list, replace_list):
            TargetText.replace(o, r)

        return TargetText

# class TrainingThread(QThread):
#     Train_end_sig = pyqtSignal(int)
#
#     def __init__(self, parent=None):
#         super ().__init__ ()
#
#     def run(self):
#
#         inputs = [mainui.trainParams.item (i, 0).text () for i in range (0, 13)]
#         parser = argparse.ArgumentParser ()
#
#
#         parser.add_argument ("--task", default="naver-ner", type=str, help="The name of the task to train")
#         parser.add_argument ("--data_dir", default='/'.join(str (mainui.trainInput.text()).split ("', '")[0].split('/')[:-1]), type=str, help="The input data dir")
#         parser.add_argument ("--model_dir",default='./model', type=str)
#         parser.add_argument ("--pred_dir", default="./preds", type=str, help="The prediction file dir")
#
#         parser.add_argument ("--train_file", default=mainui.trainInput.text(), type=str, help="Train file")
#         parser.add_argument ("--test_file", default=mainui.trainInput_3.text(), type=str, help="Test file")
#         parser.add_argument ("--label_file", default=mainui.trainInput_2.text(), type=str, help="Slot Label file")
#         parser.add_argument ("--write_pred", action="store_true", help="Write prediction during evaluation")
#
#         parser.add_argument ("--model_type", default="kobert", type=str,
#                              help="Model type selected in the list: " + ", ".join (MODEL_CLASSES.keys ()))
#
#         parser.add_argument ('--seed', type=int, default=42, help="random seed for initialization")
#         parser.add_argument ("--train_batch_size", default=int(inputs.pop(0)), type=int, help="Batch size for training.")
#         parser.add_argument ("--eval_batch_size", default=int(inputs.pop(0)), type=int, help="Batch size for evaluation.")
#         parser.add_argument ("--max_seq_len", default=int(inputs.pop(0)), type=int,
#                              help="The maximum total input sequence length after tokenization.")
#         parser.add_argument ("--learning_rate", default=float(inputs.pop(0)), type=float, help="The initial learning rate for Adam.")
#         parser.add_argument ("--num_train_epochs", default=float(inputs.pop(0)), type=float,
#                              help="Total number of training epochs to perform.")
#         parser.add_argument ("--weight_decay", default=float(inputs.pop(0)), type=float, help="Weight decay if we apply some.")
#         parser.add_argument ('--gradient_accumulation_steps', type=int, default=int(inputs.pop(0)),
#                              help="Number of updates steps to accumulate before performing a backward/update pass.")
#         parser.add_argument ("--adam_epsilon", default=float(inputs.pop(0)), type=float, help="Epsilon for Adam optimizer.")
#         parser.add_argument ("--max_grad_norm", default=float(inputs.pop(0)), type=float, help="Max gradient norm.")
#         parser.add_argument ("--max_steps", default=int(inputs.pop(0)), type=int,
#                              help="If > 0: set total number of training steps to perform. Override num_train_epochs.")
#         parser.add_argument ("--warmup_steps", default=int(inputs.pop(0)), type=int, help="Linear warmup over warmup_steps.")
#
#         parser.add_argument ('--logging_steps', type=int, default=int(inputs.pop(0)), help="Log every X updates steps.")
#         parser.add_argument ('--save_steps', type=int, default=int(inputs.pop(0)), help="Save checkpoint every X updates steps.")
#
#         parser.add_argument ("--do_train", action="store_true", help="Whether to run training.")
#         parser.add_argument ("--do_eval", action="store_true", help="Whether to run eval on the test set.")
#         parser.add_argument ("--no_cuda", action="store_true", help="Avoid using CUDA when available")
#
#         args = parser.parse_args ()
#
#
#         args.model_name_or_path = MODEL_PATH_MAP[args.model_type]
#
#         kobert_main.main(args)



class LISAUI (QMainWindow, form_class):
    def __init__(self):
        self.FThr = ProcessThread ()
        # self.TThr = TrainingThread ()

        super ().__init__ ()
        self.setupUi(self)
        # label = QLabel(self)
        # label.setPixmap(QPixmap(":/images/LINITO.png"))
        QFontDatabase.addApplicationFont("./CooperFiveOpti-Black.otf")
        self.label_6.setStyleSheet("QLabel{font-size: 55px;font-family: 'CooperFiveOpti Black';color: #3B9B74}")
        self.targetBrowse.released.connect(self.Targetfile_open)
        self.modelBrowse.released.connect(self.ModelDir_open)
        # self.tagsetOpen.released.connect(self.Dic_Editting)
        # self.visButton.released.connect(self.resultwindow)
        self.startButton.released.connect(self.NERStart)
        self.Dict_browse.released.connect (self.Dict_File_open)
        self.dict_refresh.released.connect(self.refreshing)
        self.tagsetOpen.released.connect (self.Dic_Editting)
        # self.startButton.released.connect (self.Visualize)  # start 버튼이랑 연결
        # self.setupUi (self)
        # self.concorWeb.load(QUrl.fromLocalFile (os.getcwd () + r'/startbg.html'))
        # self.Recognize_Entities.triggered.connect (self.OptUi)
        # self.Open_Evaluator.triggered.connect (self.Evaluatewindow)
        # self.TraindataSave.triggered.connect (self.BIOConverterwindow)
        # self.TrainerOpen.triggered.connect(self.Trawindow)
        # self.About_Dicora.triggered.connect (self.AboutDicora_Info)


#### 시그널 및 함수 ####

        
    ## NEW Evaluation 
    # def Evaluatewindow(self):
    #     QWidget.__init__ (self)
    #     uic.loadUi ("./UI/EVAL.ui", self)
    #
    #     self.Model_browse.released.connect (self.Eval_Model_open)
    #     self.Eval_Mode.activated.connect(self.Eval_modeboxState)
    #     self.Ans_browse.released.connect (self.Eval_Ans_open)
    #     self.Target_browse.released.connect (self.Eval_Target_open)
    #     self.evaluate_button.released.connect (self.Evaluate)
    #     self.dict_refresh.released.connect(self.refreshing)
    #     self.tagsetOpen.released.connect (self.Dic_Editting)
    #
    #     self.show()


    #
    # #### 상호 작용 함수 ####
    #
    #
    # def Eval_modeboxState(self):
    #     if self.Eval_Mode.currentText() == "Only LGG Evaluate":                 #=> combobox가 문자열이 선택되면~ currentText() == '뭐시기':
    #         self.Model_path_input.setEnabled(False)                  #=> False 가 막아놓기
    #         self.Model_browse.setEnabled(False)
    #     else:
    #         self.Model_path_input.setEnabled(True)
    #         self.Model_browse.setEnabled(True)
    #
    #
    # def Eval_Model_open(self):
    #     fname = QFileDialog.getExistingDirectory (self, 'Open Folder', '')
    #     self.Model_path_input.setText(fname)
    #
    # def Eval_Ans_open(self):
    #     fname = QFileDialog.getOpenFileName(None, 'Open Text file', '',"txt File(*.txt)")
    #     self.Ans_path_input.setText((str(fname)).split("', '")[0][2:])
    #
    # def Eval_Target_open(self):
    #     fname = QFileDialog.getOpenFileName(None, 'Open Text file', '',"txt File(*.txt)")
    #     self.Target_path_input.setText((str(fname)).split("', '")[0][2:])
    #
    # # def Eval_label_open(self):
    # #     fname = QFileDialog.getOpenFileName(None, 'Open Text file', '',"txt File(*.txt)")
    # #     self.label_path_input.setText((str(fname)).split("', '")[0][2:])
    #
    # def Evaluate(self):
    #     if self.Eval_Mode.currentText() == "Only LGG Evaluate":
    #         Eval_def.DecoEvaluator(answer_file=self.Ans_path_input.text(),
    #                                target_file=self.Target_path_input.text(),
    #                                label_path=os.getcwd () + r'/DefaultDic.xlsx',
    #                                label_sheet_name=self.Tagsetmod_combo.currentText(),
    #                                eval_mode=self.Tag_mode.currentText())
    #
    #     else:
    #         _, _ = predict (
    #             pred_model=self.Model_path_input.text(),
    #             pred_input=self.Target_path_input.text(),
    #             pred_label=os.getcwd () + r'/DefaultDic.xlsx',
    #             pred_batch_size=int(self.pred_batch.text()),
    #             label_sheet_name=self.Tagsetmod_combo.currentText())
    #
    #         Eval_def.DecoEvaluator(answer_file=self.Ans_path_input.text(),
    #                                target_file="%s_pred_out.txt" % self.Target_path_input.text()[:-4],
    #                                label_path=os.getcwd () + r'/DefaultDic.xlsx',
    #                                label_sheet_name=self.Tagsetmod_combo.currentText(),
    #                                eval_mode=self.Tag_mode.currentText())


    # ## 시각화 및 Prediction
    # def OptUi(self):
    #     QWidget.__init__ (self)
    #     uic.loadUi ("./UI/LIEN.ui", self)
    #
    #     self.targetBrowse.released.connect (self.Targetfile_open)
    #     self.modelBrowse.released.connect (self.ModelDir_open)
    #     self.tagsetOpen.released.connect (self.Dic_Editting)
    #     self.startButton.released.connect (self.NERStart)
    #     self.startButton.released.connect (self.Visualize)    # start 버튼이랑 연결
    #     # self.dict_refresh.released.connect(self.refreshing)
    #     # self.Plabel_browse.released.connect (self.Plabel_open)
    #
    #     self.show ()


    def refreshing(self):
        color_file = load_workbook(os.getcwd() + r'/DefaultDic.xlsx')
        modename = color_file.sheetnames
        self.Tagsetmod_combo.clear()
        self.Tagsetmod_combo.addItems(modename)
        color_file.close()

    # def StatUi(self):
    #     QWidget.__init__ (self)
    #     uic.loadUi ("./UI/stat_sub.ui", self)
    #
    #     self.show ()
    def resultwindow(self):
        QWidget.__init__ (self)
        uic.loadUi ("./UI/LISA_res.ui", self)

        # self.concorWeb.load(QUrl.fromLocalFile (os.getcwd () + r'/Htmls/Textvis.html'))

        self.show()
    
    # def Visualize(self):
    #     QWidget.__init__ (self)
    #     uic.loadUi ("./UI/LIEN_res.ui", self)

    #     # self.concorWeb.load(QUrl.fromLocalFile (os.getcwd () + r'/Htmls/Textvis.html'))

    #     self.show()

    #     vislist = [os.getcwd () + r'/Htmls/Textvis.html', os.getcwd () + r'/Htmls/graph.html']
    #     # self.StatUi ()
    #     vislist = [QUrl.fromLocalFile (i) for i in vislist]
    #     weblist = [self.concorWeb]
    #     for i, j in zip (weblist, vislist):
    #         i.load (j)
    #     # self.StatUi ()
    #     # self.close ()

    def Targetfile_open(self):
        fname = QFileDialog.getOpenFileName (None, 'Open Training File', '', 'TXT File (*.txt)')
        self.targetInput.setText ((str (fname)).split ("', '")[0][2:])

    def ModelDir_open(self):
        dname = QFileDialog.getExistingDirectory (self, 'Open Folder', '')
        self.modelPath.setText (dname)

    def Plabel_open(self):
        fname = QFileDialog.getOpenFileName(None, 'Open Text file', '',"txt File(*.txt)")
        self.Plabel_input.setText((str(fname)).split("', '")[0][2:])

    
    def Dict_File_open(self):
        fname = QFileDialog.getOpenFileNames(None, 'Target Location', '', 'Excel File (*.xlsx)')
        self.Dict_input.setText((str(fname)).split("', '")[0][2:].strip("'"))

    def Dic_Editting(self):
        color_path = os.getcwd () + r'/DefaultDic.xlsx'
        return file_start (color_path)

    def NERStart(self):
        self.FThr.start ()
        self.FThr.finish_sig.connect(self.processfin)

        # QWidget.__init__ (self)
        # uic.loadUi ("./UI/LIEN_res.ui", self)

        # 나중에 바꿔주기
        # self.concorWeb.load(QUrl.fromLocalFile (os.getcwd () + r'/startbg.html'))
        # self.concorWeb.load(QUrl.fromLocalFile (os.getcwd () + r'/Htmls/Textvis.html'))

        

    # ## BIO Converter
    # def BIOConverterwindow(self):
    #     QWidget.__init__ (self)
    #     uic.loadUi ("./UI/BIOconverter.ui", self)
    #     self.Testdata_check.stateChanged.connect(self.checkBoxState)
    #     self.LGG_browse.released.connect (self.LGG_open)
    #     self.convert_button.released.connect (self.Convert)
    #
    #     self.show()

    #
    # def checkBoxState(self):
    #     if self.Testdata_check.isChecked():
    #         self.ratio_num.setEnabled(True)
    #     else:
    #         self.ratio_num.setEnabled(False)
    #
    # def LGG_open(self):
    #     fname = QFileDialog.getOpenFileName(None, 'Open Text file', '',"txt File(*.txt)")
    #     self.LGG_input.setText((str(fname)).split("', '")[0][2:])
    #
    # def Convert(self):
    #     if self.Testdata_check.isChecked():
    #         convert.TrainingFomater(data=self.LGG_input.text(),emptyNum=int(self.untagged_num.text()),ratio=int(self.ratio_num.text())).process()
    #     else:
    #         convert.TrainingFomater(data=self.LGG_input.text(),emptyNum=int(self.untagged_num.text()),ratio=None).process()

    # ## Trainer
    # def Trawindow(self):
    #     QWidget.__init__ (self)
    #     uic.loadUi ("./UI/Trainer.ui", self)
    #     self.trxtBrowse.released.connect(self.Trainfile_open)
    #     self.trxtBrowse_2.released.connect(self.label_open)
    #     self.trxtBrowse_3.released.connect(self.evaldata_open)
    #     self.startTrain.released.connect(self.NERTrain)
    #
    #     self.show()
    #
    # def Logwindow(self):
    #     QWidget.__init__ (mainui)
    #     uic.loadUi ("./UI/Logs.ui", self)
    #     self.show ()
    #
    # def normalOutputWritten(self, text):
    #     cursor = self.textEdit.textCursor ()
    #     cursor.movePosition (QTextCursor.End)
    #     cursor.insertText (text)
    #     self.textEdit.setTextCursor (cursor)
    #     self.textEdit.ensureCursorVisible ()
    #
    # def Trainfile_open(self):
    #     fname = QFileDialog.getOpenFileName (None, 'Open Training File', '', 'TSV File (*.tsv)')
    #     self.trainInput.setText ((str (fname)).split ("', '")[0][2:])
    #
    # def label_open(self):
    #     fname = QFileDialog.getOpenFileName (None, 'Open Label File', '', 'txt File (*.txt)')
    #     self.trainInput_2.setText ((str (fname)).split ("', '")[0][2:])
    #
    # def evaldata_open(self):
    #     fname = QFileDialog.getOpenFileName (None, 'Open Label File', '', 'TSV File (*.tsv)')
    #     self.trainInput_3.setText ((str (fname)).split ("', '")[0][2:])
    #
    # def NERTrain(self):
    #     mainui.trainbar.setMaximum(0)
    #     mainui.trainbar.setMinimum(0)
    #     self.TThr.start()
    #     self.TThr.Train_end_sig.connect(self.Progress_end)
    #
    # def Progress_end(self):
    #     mainui.trainbar.setMaximum(2)
    #     mainui.trainbar.setMinimum(1)
    #     mainui.trainbar.setValue(2)
    #
    def processfin(self, value):
        if value == 1:
            mainui.progressBar.setMaximum (1)
            mainui.progressBar.setValue(1)
            mainui.startButton.setEnabled (True)
            # mainui.visButton.setEnabled(True)

            # vislist = [os.getcwd () + r'/Htmls/Textvis.html']
            # self.resultwindow ()
            # vislist = [QUrl.fromLocalFile (i) for i in vislist]
            # weblist = [self.concorWeb]
            # for i, j in zip (weblist, vislist):
            #     i.load (j)
            
            # # self.show ()
            # self.resultwindow ()
            # self.close ()

        else:
            mainui.startButton.setDisabled (True)
            mainui.progressBar.setMaximum (0)
            mainui.progressBar.setMinimum (0)
            mainui.progressBar.setValue (0)

    # ## AboutDICORA
    # def AboutDicora_Info(self):
    #     QWidget.__init__ (self)
    #     uic.loadUi ("./UI/DICORA_INFO.ui", self)
    #     self.show ()

if __name__ == "__main__":
    app = QApplication (sys.argv)
    mainui = LISAUI ()
    mainui.show ()
    app.exec_ ()